from selenium import webdriver

import logging
import openpyxl

# Create and Configure Logger

logging.basicConfig(filename="Batch10", format='%(asctime)s %(message)s', filemode='w')

logger = logging.getLogger()

logger.setLevel(logging.DEBUG)
filename = "C:\\Users\Dell\Desktop\DLn.xlsx"

# Open and Load Excel file by name
logger.info("OpeningWorkbook : " + filename)
workbook = openpyxl.load_workbook(filename)
logger.debug("Opened Workbook : " + filename)

# Locate and Get particular sheet
sheetName = "Sheet1"
sheet = workbook[sheetName]
logger.info(f"switched to {sheetName}")

# Read Data from sheet
try:

    print(sheet.cell(row=1, column=1).value)
    print(sheet.cell(row=1, column=3).value)
    logger.debug(sheet.cell(row=1, column=1).value + " " + sheet.cell(row=1, column=2).value)
    raise Exception
except Exception as e :
    #logger.error("Something went wrong: " + e.__str__())
    logger.error("Something went wrong: " , exc_info = e)


sheet.cell(row=9, column=5).value = "Batch10"
sheet.cell(row=9, column=3).value = "Batch_10_PassOut"

#workbook.save(sheet)
