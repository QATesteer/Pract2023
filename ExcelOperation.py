import openpyxl
import ExcelOperationBasics

# workbook = openpyxl.load_workbook("C:\\Users\Dell\Desktop\RS1.xlsx")

##locate and get perticular sheet
# Read Data from sheet

# sheet = workbook['Sheet1']
#
# print(sheet.cell(row=1,column=1).value)
# print(sheet.cell(row=1,column=3).value)

path = "C:\\Users\Dell\Desktop\DLn.xlsx"
sh = path[1]
rows = ExcelOperationBasics.getRowCount(path, 'Sheet1')
print(rows)

columns = ExcelOperationBasics.getColumnCount(path,'Sheet1')
print(columns)

ReadData = ExcelOperationBasics.readData(path,'Sheet1',rowNo=2,columnNo=2)
print(ReadData)

writedata = ExcelOperationBasics.writeData(path,'Sheet1',rowNo=11,columnNo =12,data="Hello")

# write = ExcelOperationBasics.writeData(path,'Sheet1', rowNo=9, columnNo=2,data= "Waghade")
# print(write)

#################TO Print all the data from Excel

# for r in range(1,rows+1):
#     for c in range(1,columns+1):
#         d = sh.cell(r,c)
#         print(d.value)




